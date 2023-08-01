import openpyxl
import datetime

def write_results_to_excel(results, data_dict):
    # Create a new Excel workbook and select the active sheet
    wb = openpyxl.Workbook()
    sheet = wb.active

    user_name = get_full_name(data_dict["Name"])
    table_name = data_dict["Table"]

    # Write the column headers to the Excel sheet
    column_headers = [
        "Functii",
        "PlaylistId",
        "MelodieId",
        "MembruId",
        "Title",
        "Interpret",
        "Mins",
        "Sec",
        "Denumire",
        "Utilizator",
        "DataS",
        "DataF",
        "uuid",
        "valsec",
        "valsec_Cablu",
        "valsec_Amb",
        "valsec_CP",
        "sursa",
        "domeniu",
        "PlaylistLiniiId",
        "Unicitate"
    ]
    for col_index, header in enumerate(column_headers, 1):
        sheet.cell(row=1, column=col_index, value=header)

    # Write the results to the Excel sheet
    for row_index, row in enumerate(results, 2):
        for col_index, value in enumerate(row, 2):
            sheet.cell(row=row_index, column=col_index, value=value)

    # Concatenate values of specific columns (3, 7, 8, 9, 11, 12, 18, 19) and put the result in column 21
    for row in sheet.iter_rows(min_row=2, values_only=True):
        concatenated_values = "".join(row[col_index] for col_index in [3, 7, 8, 9, 11, 12, 18, 19])
        sheet.cell(row=row_index, column=21, value=concatenated_values)
        print(sheet.cell(row=row_index, column=21, value=concatenated_values))
    
    # Save the Excel file
    wb.save(name_excel_file(table_name, user_name))

def name_excel_file(table_name, user_name):
    current_date = datetime.date.today()
    formatted_date = current_date.strftime("%d%m%Y")

    # Extract the month and year from the table name
    month_year_string = table_name.split("_")[-1]

    # Extract the first three characters as the month
    month_string = month_year_string[:3].lower()

    # Extract the last four characters as the year
    year_string = month_year_string[-4:]

    final_name = f"{user_name}_exx__{month_string}{year_string}__________vali_{formatted_date}.xlsx"

    return final_name

def get_full_name(name_data):
    name_parts = [name_data[key] for key in name_data if name_data[key]]
    return " ".join(name_parts)


