import openpyxl
import datetime
import os
import subprocess

COLUMN_HEADERS = [
    "Functii", "PlaylistId", "MelodieId", "MembruId", "Title", "Interpret", "Mins",
    "Sec", "Denumire", "Utilizator", "DataS", "DataF", "uuid", "valsec", "valsec_Cablu",
    "valsec_Amb", "valsec_CP", "sursa", "domeniu", "PlaylistLiniiId", "Unicitate"
]

CONCATENATE_COLUMNS = [3, 7, 8, 9, 11, 12, 18, 19]

class ExcelProcessor:
    @staticmethod
    def create_excel_writer():
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        return workbook, sheet

    @staticmethod
    def save_excel_file(workbook, file_name):
        workbook.save(file_name)

    @staticmethod
    def get_full_name(name_data):
        name_parts = [name_data[key] for key in name_data if name_data[key]]
        return " ".join(name_parts)

    @staticmethod
    def define_sheet_structure(sheet):
        sheet.insert_cols(1)
        for col_index, header in enumerate(COLUMN_HEADERS, 1):
            sheet.cell(row=1, column=col_index, value=header)

    @staticmethod
    def write_results(sheet, results):
        for row_index, row in enumerate(results, 2):
            for col_index, value in enumerate(row, 2):
                if isinstance(value, datetime.date):
                    sheet.cell(row=row_index, column=col_index, value=value.strftime("%Y-%m-%d"))
                else:
                    sheet.cell(row=row_index, column=col_index, value=value)

    @staticmethod
    def concatenate_data(sheet):
        for row in sheet.iter_rows(min_row=2):
            concatenated_strings = ''.join(str(row[col_index - 1].value) for col_index in CONCATENATE_COLUMNS)
            sheet.cell(row=row[0].row, column=21, value=concatenated_strings)

    @staticmethod
    def get_values_with_rows(sheet):
        values_with_rows = {}
        for row_number, row in enumerate(sheet.iter_rows(min_row=2, min_col=21, values_only=True), start=2):
            cell_value = row[0]
            if cell_value not in values_with_rows:
                values_with_rows[cell_value] = [row_number]
            else:
                values_with_rows[cell_value].append(row_number)
        return values_with_rows

    @staticmethod
    def mark_duplicates(values_with_rows, sheet):
        for rows in values_with_rows.values():
            if len(rows) >= 2:
                for row in rows[1:]:
                    sheet.cell(row=row, column=1, value="Duplicat")

        for rows in values_with_rows.values():
            if len(rows) >= 2:
                for row in rows[1:]:
                    sheet.cell(row=row, column=1, value="Duplicat")
    
    @staticmethod
    def process_search_results(results, data_dict):
        current_date = datetime.date.today()
        formatted_date = current_date.strftime("%d%m%Y")

        # Extract month and year from the table name
        month_string, year_string = data_dict["Table"].split("_")[-1][:3].lower(), data_dict["Table"].split("_")[-1][-4:]

        # Extract the user's name from the data dictionary and create the file name
        user_name = ExcelProcessor.get_full_name(data_dict["Name"])
        file_name = f"{user_name}_exx__{month_string}{year_string}__________vali_{formatted_date}.xlsx"

        workbook, sheet = ExcelProcessor.create_excel_writer()
        ExcelProcessor.define_sheet_structure(sheet)
        ExcelProcessor.write_results(sheet, results)
        ExcelProcessor.concatenate_data(sheet)
        values_with_rows = ExcelProcessor.get_values_with_rows(sheet)
        ExcelProcessor.mark_duplicates(values_with_rows, sheet)
        ExcelProcessor.save_excel_file(workbook, file_name)

        print(f"Excel file '{file_name}' saved successfully.")
    
    @staticmethod
    def open_excel_file(file_name):
        # Use subprocess to open the Excel file with the default application
        try:
            subprocess.run([file_name], shell=True)
        except Exception as e:
            error_message = "An error occurred while opening the Excel file: " + str(e)
            print(error_message)
            QMessageBox.critical(None, "Error", error_message)